import numpy as np
import math
from PyQt5.QtWidgets import QAction, QApplication, QFileDialog, QMessageBox
from PyQt5 import QtWidgets, QtGui
import Herst_gui #file with gui
import matplotlib
matplotlib.use('QT5Agg')
import sys
import pyqtgraph as pg
from openpyxl import load_workbook # for import data from Excel
import openpyxl

class RS_analysis:

    def __init__(self, data_y, tau=1, del_tau=1):
        self.tau = tau #начальная точка
        self.del_tau = del_tau #шаг
        self.data_y = data_y
        self.res_g = []   #Результат построения степенной регрессии
        self.x = []   #Значения tau
        self.new_y = []   #Результат вычисления
        self.a = 0
        self.b = 0  #Показатель Херста
        self.analysis()


    def rrange(self, tau, val): # Вычисление размаха для текущего tau
        s = 0
        ave = sum(val)/tau
        min = val[0]-ave
        max = val[0]-ave
        for i in val:
            s += i - ave
            if s < min: min = s
            if s > max: max = s
        return max-min


    def SCO(self, tau, val):
        medium = sum(val)/tau
        s = 0
        for i in val:
            s += (i - medium)**2
        return (s / (tau - 0.99999999))**0.5


    def Fx(self, x):
        return self.a*(x**self.b)


    # Метод наименьших квадратов
    # least square method

    def LS(self, x, y):
        sX=0
        sY = 0
        sXY = 0
        sXX = 0
        n = len(y)

        for i in range(len(x)):
            sX += math.log(x[i])
            sXX += math.log(x[i])**2
            if (y[i]>0):
                sY += math.log(y[i])
                sXY += math.log(x[i]) * math.log(y[i])

        b = (n * sXY - sX * sY) / (n * sXX - sX**2)
        a = math.exp((sY - b * sX) / n)
        return [a, b]


    def analysis(self):
        self.new_y = []
        for i in np.arange(1, len(self.data_y), self.del_tau):
            def_R = self.rrange(i, self.data_y[:i])
            def_S = self.SCO(i, self.data_y[:i])
            if math.isnan(def_R / def_S):
                self.new_y.append(0)
            else:
                self.new_y.append(def_R / def_S)
            self.x.append(i)
        self.new_y = np.array(self.new_y)
        self.a = self.LS(self.x, self.new_y)[0]
        self.b = self.LS(self.x, self.new_y)[1]
        for i in self.x:
            self.res_g.append(self.Fx(i))
            
    #tuple for import result in txt/excel file
    def doTuple(self):
        result_tuple = []
        result_tuple.append(['a', self.a])
        result_tuple.append(['b', self.b])
        result_tuple.append(['x', 'R/S', 'Степенная регрессия'])
        for i in range(len(self.new_y)):
            result_tuple.append([self.x[i], self.new_y[i], self.res_g[i]])
        return result_tuple

class Herst_window(QtWidgets.QMainWindow, Herst_gui.Ui_MainWindow):
    def __init__(self, *args, **kwargs):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle('Показатель Херста')
        self.__datafile = []
        self.__result_tuple = []
        self.lineEdit.setText('1')
        self.lineEdit_2.setText('1')
        self.lineEdit.setValidator(QtGui.QIntValidator())
        self.lineEdit_2.setValidator(QtGui.QIntValidator())
        self.pushButton.setEnabled(0)

        openFile = QAction('Открыть файл', self)
        openFile.setShortcut('Ctrl+O')
        openFile.setStatusTip('Открыть файл')
        openFile.triggered.connect(self.showDialog)
        self.menu.addAction(openFile)

        self.GraphWidget.setBackground('w')
        style = {'color': 'b', 'font-size': '20px'}
        self.GraphWidget.setLabel('left', 'R/S', **style)
        self.GraphWidget.setLabel('bottom', 'time', **style)
        self.GraphWidget.addLegend()

        saveData = self.menu_2.addAction("Сохранить данные в Excel")
        saveData.triggered.connect(self.saveData)

        showOriginData = self.menu_2.addAction("Показать исходные данные")
        showOriginData.triggered.connect(self.showData)

    def set_herst_num(self, value='-'):
        self.herst_num.setText(value)

    def set_frac(self, value='-'):
        self.frac.setText(value)

    def set_mandelb(self, value='-'):
        self.mandelb.setText(value)

    def set_corr(self, value='-'):
        self.corr.setText(value)

    def set_plot(self, x, y, pen, lbl='-'):
        self.GraphWidget.plot(x, y, pen=pen, name=lbl)

    def set_resultTuple(self, tup):
        self.__result_tuple = tup

    def get_tau(self):
        return int(self.lineEdit.text())

    def get_delta_tau(self):
        return int(self.lineEdit_2.text())

    def get_datafile(self):
        return self.__datafile

    def clear_plot(self):
        self.GraphWidget.clear()
        # print(self.GraphWidget.plotItem)

    def showDialog(self):
        try:
            fname = QFileDialog.getOpenFileName(self, "Открыть текстовый файл", "/home", "Excel (*.xlsx);;Data (*.dat);; Text (*.txt)")[0]
            if (fname.find('.dat')!=-1 or fname.find('.txt')!=-1):
                fromfile = open(fname, 'r')
                lines = fromfile.readlines()
                self.__datafile = [float(i) for i in lines]
                fromfile.close()
            else:
                fromfile = load_workbook(fname)['RC']
                sheet_range = fromfile['A']
                for cell in sheet_range:
                    self.__datafile.append(float(cell.value))
        except:
            self.showWindowMsg(QMessageBox.Warning, 'Ошибка!', 'Неверные данные, проверьте файл')

        if len(self.__datafile) != 0:
            self.pushButton.setEnabled(1)


    def saveData(self):
        try:
            fname = QFileDialog.getSaveFileName(self, "Сохранить файл данных", "/home", "Excel (*.xlsx)")[0]
            wb = openpyxl.Workbook()
            wb.create_sheet(title='Результат', index=0)
            sheet = wb['Результат']
            for i in range(len(self.__result_tuple)):
                sheet.append(self.__result_tuple[i])
            wb.save(fname)
            self.showWindowMsg(QMessageBox.Information, 'Результат', 'Данные успешно сохранены')
        except:
            self.showWindowMsg(QMessageBox.Warning, 'Ошибка!', 'Проблемы при сохранении файла')

    def showData(self):
        self.GraphWidget.clear()
        style = {'color': 'b', 'font-size': '20px'}
        self.GraphWidget.setLabel('left', 'r(t)', **style)
        x = [i for i in range(len(self.__datafile))]
        pen = pg.mkPen(color=(1, 88, 239), width=5)
        self.set_plot(x, self.datafile, pen, 'Исходные данные')

    def showWindowMsg(self, type, title, text):
        msg = QMessageBox()
        msg.setIcon(type)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()


def visualization(self):
    tau = int(window.lineEdit.text())
    del_tau = int(window.lineEdit_2.text())
    RS = RS_analysis(window.get_datafile(), tau, del_tau)
    window.set_resultTuple(RS.doTuple())
    window.set_herst_num(str(round(RS.b, 4)))
    window.set_mandelb(str(round(1/RS.b, 4)))
    window.set_frac(str(round(2-RS.b, 4)))
    window.set_corr(str(round(2**(2*RS.b-1)-1, 4)))
    window.clear_plot()
    pen = pg.mkPen(color=(1, 88, 239), width=5)
    window.set_plot(RS.x, RS.new_y, pen, 'Обработанные данные')
    pen = pg.mkPen(color=(255, 191, 0), width=5)
    window.set_plot(RS.x, RS.res_g, pen, 'Регрессия y=a*x^b')

app = QApplication(sys.argv)
window = Herst_window()
window.pushButton.clicked.connect(visualization)
window.show()
app.exec_()

